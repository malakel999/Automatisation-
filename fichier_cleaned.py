import pandas as pd
from openpyxl import load_workbook

# Fonction pour lire un fichier Excel volumineux avec openpyxl par lots
def lire_lignes_excel_par_lots(fichier, taille_lot=100):
    try:
        # Charger le fichier Excel en mode lecture seule
        wb = load_workbook(fichier, read_only=True)
        print(f"\nAnalyse du fichier : {fichier}")
        
        # Parcourir chaque feuille
        for feuille in wb.sheetnames:
            print(f"\nAperçu de la feuille '{feuille}':")
            ws = wb[feuille]
            start_row = 0  # Initialiser le compteur de lignes

            # Lire et traiter le fichier par lots
            while True:
                lignes = []
                # Lire un lot de lignes
                for i, row in enumerate(ws.iter_rows(min_row=start_row + 1, max_row=start_row + taille_lot, values_only=True), start=1):
                    lignes.append(row)
                    print(row)  # Afficher la ligne

                if not lignes:
                    break  # Terminer si plus de données

                # Afficher le nombre de lignes lues
                print(f"Lignes lues de {start_row + 1} à {start_row + len(lignes)} dans '{feuille}'")

                # Passer au prochain lot
                start_row += taille_lot

    except Exception as e:
        print(f"Erreur lors de la lecture du fichier {fichier}: {e}")

# Fonction pour analyser des fichiers Excel volumineux avec pandas par lots
def analyse_fichier_excel_large_par_lots(fichier, taille_lot=100):
    try:
        # Charger le fichier Excel avec pandas
        xls = pd.ExcelFile(fichier)
        print(f"\nAnalyse du fichier : {fichier}")
        print(f"Feuilles présentes : {xls.sheet_names}")
        
        # Parcourir chaque feuille
        for feuille in xls.sheet_names:
            print(f"\nAperçu de la feuille '{feuille}':")
            start_row = 0  # Initialiser le compteur de lignes

            # Lire et traiter le fichier par lots
            while True:
                # Lire un lot de données
                df_lot = pd.read_excel(fichier, sheet_name=feuille, skiprows=start_row, nrows=taille_lot)
                if df_lot.empty:
                    break  # Terminer si plus de données

                # Afficher les colonnes et un aperçu des données
                print(f"Colonnes : {df_lot.columns}")
                print("Aperçu des lignes lues :")
                print(df_lot.head())  # Afficher les 5 premières lignes du lot

                # Afficher le nombre de lignes chargées
                print(f"Lignes lues de {start_row + 1} à {start_row + len(df_lot)} dans '{feuille}'")

                # Passer au prochain lot
                start_row += taille_lot

    except Exception as e:
        print(f"Erreur lors de la lecture du fichier {fichier}: {e}")

# Chemins des fichiers Excel
fichier_1 = 'Ventes et avoirs mi juillet 2024.xlsx'
fichier_2 = 'Factures PR MFS 2024 4 check du 24 07 2024 (1).xlsx'

# Utiliser openpyxl pour lire les fichiers Excel par lots
print("Lecture avec openpyxl par lots :")
lire_lignes_excel_par_lots(fichier_1, taille_lot=100)
lire_lignes_excel_par_lots(fichier_2, taille_lot=100)

# Utiliser pandas pour analyser les fichiers Excel par lots
print("\nLecture avec pandas par lots :")
analyse_fichier_excel_large_par_lots(fichier_1, taille_lot=100)
analyse_fichier_excel_large_par_lots(fichier_2, taille_lot=100)
