import pandas as pd
from openpyxl import load_workbook

# Fonction principale pour lire, nettoyer et générer un nouveau fichier Excel
def traiter_fichier_excel(fichier, methode='openpyxl', taille_lot=100):
    print(f"\n--- Traitement du fichier : {fichier} ---")

    # Étape 1 : Lecture du fichier Excel
    if methode == 'openpyxl':
        donnees_nettoyees = lire_excel_avec_openpyxl(fichier, taille_lot)
    elif methode == 'pandas':
        donnees_nettoyees = lire_excel_avec_pandas(fichier, taille_lot)
    else:
        print(f"Méthode '{methode}' non reconnue. Utilisez 'openpyxl' ou 'pandas'.")
        return

    # Étape 2 : Nettoyage des données (déjà effectué lors de la lecture)
    # Les données sont déjà nettoyées dans les fonctions 'lire_excel_avec_openpyxl' et 'lire_excel_avec_pandas'

    # Étape 3 : Génération d'un nouveau fichier Excel avec les données nettoyées
    generer_fichier_excel(donnees_nettoyees, fichier, methode)

# Étape 1 : Fonction pour lire et nettoyer le fichier Excel avec openpyxl
def lire_excel_avec_openpyxl(fichier, taille_lot=100):
    try:
        # Charger le fichier Excel en mode lecture seule
        wb = load_workbook(fichier, read_only=True)
        print(f"\nLecture du fichier avec openpyxl : {fichier}")
        
        # Dictionnaire pour stocker les données nettoyées par feuille
        donnees_nettoyees = {}

        # Parcourir chaque feuille
        for feuille in wb.sheetnames:
            print(f"\nTraitement de la feuille '{feuille}':")
            ws = wb[feuille]
            start_row = 0
            data_clean_cumulative = []

            # Lecture par lots
            while True:
                # Lire un lot de lignes
                lignes = []
                for i, row in enumerate(ws.iter_rows(min_row=start_row + 1, max_row=start_row + taille_lot, values_only=True), start=1):
                    # Filtrer les cellules vides
                    ligne_propre = [cell for cell in row if cell is not None]
                    if ligne_propre:  # Ajouter la ligne si elle n'est pas vide
                        lignes.append(ligne_propre)

                if not lignes:  # Si le lot est vide, terminer la boucle
                    break
                
                # Ajouter les données nettoyées au lot cumulatif
                data_clean_cumulative.extend(lignes)

                # Passer au prochain lot
                start_row += taille_lot

            # Stocker les données nettoyées pour cette feuille
            donnees_nettoyees[feuille] = data_clean_cumulative
        
        return donnees_nettoyees

    except Exception as e:
        print(f"Erreur lors de la lecture du fichier {fichier} avec openpyxl: {e}")
        return {}

# Étape 1 : Fonction pour lire et nettoyer le fichier Excel avec pandas
def lire_excel_avec_pandas(fichier, taille_lot=100):
    try:
        # Charger le fichier Excel avec pandas
        xls = pd.ExcelFile(fichier)
        print(f"\nLecture du fichier avec pandas : {fichier}")
        
        # Dictionnaire pour stocker les données nettoyées par feuille
        donnees_nettoyees = {}

        # Parcourir chaque feuille
        for feuille in xls.sheet_names:
            print(f"\nTraitement de la feuille '{feuille}':")
            start_row = 0  # Début du lot
            data_clean_cumulative = pd.DataFrame()

            while True:
                # Lire le prochain lot de données
                df_lot = pd.read_excel(fichier, sheet_name=feuille, skiprows=start_row, nrows=taille_lot)
                if df_lot.empty:
                    break  # Arrêter si toutes les lignes ont été traitées
                
                # Nettoyer les cellules vides
                df_lot_clean = df_lot.dropna(axis=1, how='all').dropna(axis=0, how='all')

                # Ajouter les données nettoyées au DataFrame cumulatif
                data_clean_cumulative = pd.concat([data_clean_cumulative, df_lot_clean], ignore_index=True)

                # Passer au prochain lot
                start_row += taille_lot

            # Stocker les données nettoyées pour cette feuille
            donnees_nettoyees[feuille] = data_clean_cumulative
        
        return donnees_nettoyees

    except Exception as e:
        print(f"Erreur lors de la lecture du fichier {fichier} avec pandas: {e}")
        return {}

# Étape 3 : Fonction pour générer un nouveau fichier Excel avec les données nettoyées
def generer_fichier_excel(donnees_nettoyees, fichier_original, methode):
    try:
        # Générer un nom pour le nouveau fichier
        suffixe = "_openpyxl" if methode == 'openpyxl' else "_pandas"
        nom_nouveau_fichier = f'nettoye_{fichier_original.split(".")[0]}{suffixe}.xlsx'

        # Enregistrer les données nettoyées dans un nouveau fichier Excel
        with pd.ExcelWriter(nom_nouveau_fichier) as writer:
            for feuille, data in donnees_nettoyees.items():
                if isinstance(data, pd.DataFrame):
                    df = data
                else:
                    df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=feuille, index=False)
        print(f"Les données nettoyées ont été enregistrées dans le fichier : {nom_nouveau_fichier}")
    
    except Exception as e:
        print(f"Erreur lors de la génération du fichier Excel nettoyé : {e}")

# Chemins des fichiers Excel
fichier_1 = 'Ventes et avoirs mi juillet 2024.xlsx'
fichier_2 = 'Factures PR MFS 2024 4 check du 24 07 2024 (1).xlsx'

# Traiter les fichiers Excel
traiter_fichier_excel(fichier_1, methode='openpyxl', taille_lot=100)
traiter_fichier_excel(fichier_2, methode='pandas', taille_lot=100)
